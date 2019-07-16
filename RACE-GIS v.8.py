#Stephen Duncanon
#RACE GIS 
#7/12/19 -

#Bugs
#if there is a space after any of the database fields it wont be picked up
#projects with blank fields will either be ommitted or included
#which compromises the search
#if it lacks coordinates, it will not be in spec loc search
#if there is any specification, None types shall be removed

#Imports
import time
import os
import os.path
import tkinter as tk
from tkinter import ttk
import datetime
import time
import openpyxl
from geopy.geocoders import Nominatim
from geopy.distance import geodesic

#Constants
VERSION = .8                                    
PROJECT_SPREADSHEET = 'X:/RACEGIS/projects.xlsx'
CLIENT_SPREADSHEET = 'X:/RACEGIS/clients.xlsx'
DATABASE_SPREADSHEET = 'X:/RACEGIS/database.xlsx'
PROJECT_SHEET_NAME = 'RACEGIS_Projects'     #This will be the name of the AJERA custom inquiry
CLIENT_SHEET_NAME = 'RACEGIS_Clients'       #This will be the name of the AJERA custom inquiry
DATABASE_SHEET_NAME = 'data'

#Variables
now = datetime.datetime.now()
geolocator = Nominatim(timeout=10,user_agent="RACE-GIS")

#KML text
kml_header_1='<?xml version="1.0" encoding="UTF-8"?>'
kml_header =\
'''<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>RACE GIS</name>
  <Style id="project-style">
  <LabelStyle>
   <scale>0</scale>
  </LabelStyle>
   <IconStyle>
    <scale>1</scale>
    <Icon>
      <href>http://maps.google.com/mapfiles/kml/paddle/blu-blank.png</href>
    </Icon>
    </IconStyle>
    <BalloonStyle>
      <text>
        <![CDATA[
	  <h2>$[name]</h2>
          <p><b>ID: </b>$[ID]</p>
	  <p><b>Address: </b>$[Address]</p>
          <p><b>Project Type: </b>$[projectType]</p>
          <p><b>Client Name: </b>$[clientName]</p>
          <p><b>Client Type: </b>$[clientType]</p>
        ]]>
      </text>
    </BalloonStyle>
  </Style>
  <!-- Shared style sample
        Two Placemarks use the same balloon template
  -->
'''


kml_footer=\
'''</Document>
</kml>'''

#Functions
def gen_placemark(name,proj_id,location,proj_type,client,client_type,latlongstr):

    placemark = """
  <Placemark>
    <name>"""+name+"""</name>
    <styleUrl>#project-style</styleUrl>
    <ExtendedData>
      <Data name="ID">
	<value>"""+proj_id+"""</value>
      </Data>
      <Data name="Address">
        <value>"""+location+"""</value>
      </Data>
      <Data name="projectType">
        <value>"""+prof_type+"""</value>
      </Data>
      <Data name="clientType">
        <value>"""+clien_typet+"""</value>
      </Data>
      <Data name="clientName">
        <value>"""+client+"""</value>
      </Data>
    </ExtendedData>
    <Point>
         <coordinates>"""+latlongstr+""",0</coordinates>    
    </Point>
  </Placemark>
    """


def check_for_spreadsheets():
    '''
    The initial check for project and database spreadsheets
    '''
    if os.path.exists(PROJECT_SPREADSHEET) and os.path.exists(CLIENT_SPREADSHEET):
        return True
    else:
        map_button.config(state=tk.DISABLED)
        excel_button.config(state=tk.DISABLED)
        if not os.path.exists(PROJECT_SPREADSHEET):
            popup("Project spreadsheet is missing!\n Please ensure that the following file is present:\n"+str(PROJECT_SPREADSHEET))
        if not os.path.exists(CLIENT_SPREADSHEET):
            popup("Client Spreadsheet is missing!\n Please ensure that the following file is present:\n"+str(CLIENT_SPREADSHEET))
    return False


def check_database():
    '''
    '''
    if os.path.exists(DATABASE_SPREADSHEET):
        #if it exists, check if it is updated
        if is_database_updated():
            return True
        #if it is not updated, update it
        else:
            update_database()
    else:
        generate_database()


def is_database_updated():
    project_wb = openpyxl.load_workbook(PROJECT_SPREADSHEET)
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)

    project_sheet = project_wb[PROJECT_SHEET_NAME]
    database_sheet = database_wb[DATABASE_SHEET_NAME]

        #Check how many rows are in each spreadsheet

    rows_in_project_sheet = int(project_sheet.max_row)
    rows_in_database_sheet = int(database_sheet.max_row)

    if rows_in_project_sheet == rows_in_database_sheet:
        print("The database is up to date!")
        return True
    elif rows_in_project_sheet > rows_in_database_sheet:
        print("FALSE. Database needs to be updated")
        return False
    else:
       popup("Error, database is ahead of project source, outdated project files?")

    
        
def popup(msg):
    popup = tk.Tk()
    popup.wm_title("ERROR")
    label = ttk.Label(popup, text=msg)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Close", command = popup.destroy)
    B1.pack()
    popup.mainloop()    
    

def generate_database():
    '''
    if database is missing
    '''
    print("database gen")
    #popup("Generating new database!\n This may take awhile...")
    create_database()
    
    

def create_database():
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0,title=DATABASE_SHEET_NAME)
    database_sheet = wb[DATABASE_SHEET_NAME]
    wb.remove(wb['Sheet'])
    database_sheet["A1"].value="Project ID"
    database_sheet["B1"].value="Project Name"
    database_sheet["C1"].value="Location"
    database_sheet["D1"].value="Project Type"
    database_sheet["E1"].value="Client"
    database_sheet["F1"].value="Client Type"
    database_sheet["G1"].value="Lat"
    database_sheet["H1"].value="Long"
    database_sheet["I1"].value="Permit"
    database_sheet["J1"].value="Boring"
    wb.save(DATABASE_SPREADSHEET)
    print("Created Database")
    update_database()

def merge_clients_projects():
    '''
    '''
    print("Merging databases!")
    client_wb = openpyxl.load_workbook(CLIENT_SPREADSHEET)
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)

    client_sheet = client_wb[CLIENT_SHEET_NAME]
    database_sheet = database_wb[DATABASE_SHEET_NAME]

    #create dictionary with key value pairs for all the clients, Much faster!
    #good job stephen
    d= {}
    for x in range(2,client_sheet.max_row+1):
        #dictionary["Key string"] = "value"
        d[client_sheet["A"+str(x)].value] = str(client_sheet["B"+str(x)].value)
        
    for i in range(2,database_sheet.max_row+1):
        #if database_sheet["F"+str(i)].value == "":
        database_sheet["F"+str(i)].value= d[database_sheet["E"+str(i)].value]

      
        #change to only look at blank rows
        #else:
           # break
    database_wb.save(DATABASE_SPREADSHEET)
    geocode_database()
    


def geocode_database():
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)
    database_sheet = database_wb[DATABASE_SHEET_NAME]
    print("bulk geocode with nominatim, this WILL take awhile.")
    for l in range(2,database_sheet.max_row+1):
        print(l)
        if database_sheet["G"+str(l)].value == None:
            time.sleep(1)#respect for Nominatim
            try:
                loc = geolocator.geocode(database_sheet["C"+str(l)].value)
                lat = loc.latitude
                long= loc.longitude
                database_sheet["G"+str(l)].value = lat
                database_sheet["H"+str(l)].value = long
            except:
                AttributeError
    print("added!")
    database_wb.save(DATABASE_SPREADSHEET)



def update_database():
    '''
    if the database is incomplete, this fuction is a slow version
    if I wanted to make this fast I could have it only look for new information
    but for the sake of ruggedness I think I will make it do a full comparison between the two sheets 
    '''
    ids_in_project_sheet = []
    
    project_wb = openpyxl.load_workbook(PROJECT_SPREADSHEET)
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)

    project_sheet = project_wb[PROJECT_SHEET_NAME]
    database_sheet = database_wb[DATABASE_SHEET_NAME]
    
    #maybe to speed things up this could get the data from the last function?
    #an even faster way could be to create a 2d block around new items and bulk importing them
    current_database_row = int(database_sheet.max_row)#current write row is at bottom of spreadsheet
    current_project_row = int(project_sheet.max_row)

    #Get a list of all the ids in the project file
    for i in range(2,project_sheet.max_row+1):
        ids_in_project_sheet.append(project_sheet["A"+str(i)].value)

    for j in range(2,database_sheet.max_row+1):
        try:
            ids_in_project_sheet.remove(database_sheet["A"+str(j)].value)
        except:
            ValueError

    for k in range(2,project_sheet.max_row+1):
        if str(project_sheet["A"+str(k)].value) in str(ids_in_project_sheet):
            print(project_sheet["A"+str(k)].value,"in ids!")
            #then it is a job which needs to be added
            database_sheet["A"+str(current_database_row)].value=project_sheet["A"+str(k)].value
            database_sheet["B"+str(current_database_row)].value=project_sheet["B"+str(k)].value
            database_sheet["C"+str(current_database_row)].value=project_sheet["C"+str(k)].value
            database_sheet["D"+str(current_database_row)].value=project_sheet["D"+str(k)].value
            database_sheet["E"+str(current_database_row)].value=project_sheet["E"+str(k)].value
            #PERMITS AND BORING
            #database_sheet["G"+str(current_database_row)].value=project_sheet["G"+str(k)].value
            #database_sheet["H"+str(current_database_row)].value=project_sheet["H"+str(k)].value          
            current_database_row+=1
        else:
            print(project_sheet["A"+str(k)].value,"not in ids!")

    print("added sheets from project file")
    database_wb.save(DATABASE_SPREADSHEET)
    merge_clients_projects()
    

def map_gen():
    '''
    When map button is pressed
    '''
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)
    database_sheet = database_wb[DATABASE_SHEET_NAME]
    coords = get_location()

    projects = get_project_types()
    years = get_years() #list of years to select
    clients = get_client_types() #list of client types

    #check if location is checked
    if loc_spec_var.get():
        #if it is get the location to select by
        coords = get_location()
        radius = int(loc_radius_var.get())
    export_file = open('RACEGIS.kml','w')#open a text file of the date + rest
    export_file.write(kml_header_1)
    export_file.write(kml_header)
    for i in range(2,database_sheet.max_row):
        #FILTER
        proj_id = database_sheet["A"+str(i)].value
        proj_year = str(proj_id)[:4]
        if int(proj_year) in years:#check if meets year criteria
            if database_sheet["D"+str(i)].value in projects:#check if meets project type criteria
                if database_sheet["F"+str(i)].value in clients:#check if meets client type criteria
                    if not loc_spec_var.get():#if we do NOT need to sort by location
                        export_file.write('  <Placemark>\n') 
                        export_file.write('    <name>'+database_sheet["B"+str(i)].value+'</name>\n')
                        export_file.write('    <styleUrl>#project-style</styleUrl>\n')
                        export_file.write('    <ExtendedData>\n')
                        export_file.write('      <Data name="ID">\n')
                        export_file.write('	<value>'+str(database_sheet["A"+str(i)].value)+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('      <Data name="Address">\n')
                        try:
                            export_file.write('        <value>'+database_sheet["C"+str(i)].value+'</value>\n')
                            export_file.write('      </Data>\n')
                            export_file.write('      <Data name="projectType">\n')
                            export_file.write('        <value>'+database_sheet["D"+str(i)].value+'</value>\n')
                            export_file.write('      </Data>\n')
                            export_file.write('      <Data name="clientType">\n')
                            export_file.write('        <value>'+database_sheet["F"+str(i)].value+'</value>\n')
                            export_file.write('      </Data>\n')
                            export_file.write('      <Data name="clientName">\n')
                            export_file.write('        <value>'+database_sheet["E"+str(i)].value+'</value>\n')
                            export_file.write('      </Data>\n')
                            export_file.write('    </ExtendedData>\n')
                            export_file.write('    <Point>\n')
                            try:
                                export_file.write('         <coordinates>'+str(database_sheet["H"+str(i)].value)+','+str(database_sheet["G"+str(i)].value)+',0</coordinates>\n')
                                export_file.write('    </Point>\n')
                                export_file.write('  </Placemark>\n')#why are lat and long switched??
                            except:
                                TypeError
                        except:
                            TypeError
                    else:
                        if database_sheet["G"+str(i)].value != None:
                            t_coords = (float(database_sheet["G"+str(i)].value),float(database_sheet["H"+str(i)].value))
                            distance = geodesic(coords,t_coords).miles
                            print(distance)
                            if distance < radius:
                                export_file.write('  <Placemark>\n') 
                                export_file.write('    <name>'+database_sheet["B"+str(i)].value+'</name>\n')
                                export_file.write('    <styleUrl>#project-style</styleUrl>\n')
                                export_file.write('    <ExtendedData>\n')
                                export_file.write('      <Data name="ID">\n')
                                export_file.write('	<value>'+str(database_sheet["A"+str(i)].value)+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="Address">\n')
                                export_file.write('        <value>'+database_sheet["C"+str(i)].value+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="projectType">\n')
                                export_file.write('        <value>'+database_sheet["D"+str(i)].value+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="clientType">\n')
                                export_file.write('        <value>'+database_sheet["F"+str(i)].value+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="clientName">\n')
                                export_file.write('        <value>'+database_sheet["E"+str(i)].value+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('    </ExtendedData>\n')
                                export_file.write('    <Point>\n')
                                try:
                                    export_file.write('         <coordinates>'+str(database_sheet["H"+str(i)].value)+','+str(database_sheet["G"+str(i)].value)+',0</coordinates>\n')
                                    export_file.write('    </Point>\n')
                                    export_file.write('  </Placemark>\n')#why are lat and long switched??
                                except:
                                    TypeError
    export_file.write(kml_footer)
    export_file.close()
    os.startfile("RACEGIS.kml")


def get_years():
    years = []
    for x in range(int(year_var_1.get()),int(year_var_2.get())+1):
        years.append(x)
    return years

def get_client_types():
    client_types = ['Residential',
                    'Government',
                    'Marina / Yacht Club',
                    'Industrial / Utility',
                    'Consultant',
                    'Contractor',
                    'None']#None is there, source of ERROR
    if all_client_var.get():
        return client_types
    elif not residential_var.get() and \
         not government_var.get() and \
         not myc_var.get() and \
         not util_var.get() and \
         not consultant_var.get() and \
         not contractor_var.get():
            print("ERROR! SELECT a client type")
    else:
        if not residential_var.get():
            client_types.remove('Residential')
        if not government_var.get():
            client_types.remove('Government')
        if not myc_var.get():
            client_types.remove('Marina / Yacht Club')
        if not util_var.get():
            client_types.remove('Industrial / Utility')
        if not consultant_var.get():
            client_types.remove('Consultant')
        if not contractor_var.get():
            client_types.remove('Contractor')
        client_types.remove('None')#if there is any spec, remove None.
        return client_types

def get_project_types():
    project_types = ['Dredging',#0
                     'Flood and erosion control structure',#1
                     'Beach / Dune / Living Shoreline',#2
                     'Pier / Dock',#3
                     'Hydrographic Survey',#4
                     'Facility Assessment',#5
                     'Construction Administration',#6
                     'Flood / Wave Study',#7
                     'Upland Structure',#8
                     'Other',#9
                     'None'#Source of error
                     ]
    #check each checkbox, if not checked, remove from project types list
    #print list of projects that are wanted
    #first check if all is cheched
    if all_proj_var.get():
        return project_types
        #this case quickly sees if all projects are checked
        #since if all is checked they all must be?
    elif not dredge_var.get() and \
         not floodwave_var.get() and \
         not hydro_var.get() and \
         not pierdock_var.get() and \
         not floodcontrol_var.get() and \
         not upland_var.get() and \
         not facilities_var.get() and \
         not beach_var.get() and \
         not construction_var.get() and \
         not other_var.get():
            print("ERROR! SELECT a project type")
    else:
        if not dredge_var.get():
            project_types.remove('Dredging')
        if not floodwave_var.get():
            project_types.remove('Flood / Wave Study')
        if not hydro_var.get():
            project_types.remove('Hydrographic Survey')
        if not pierdock_var.get():
            project_types.remove('Pier / Dock')
        if not floodcontrol_var.get():
            project_types.remove('Flood and erosion control structure')
        if not upland_var.get():
            project_types.remove('Upland Structure')
        if not facilities_var.get():
            project_types.remove('Facility Assessment')
        if not beach_var.get():
            project_types.remove('Beach / Dune / Living Shoreline')
        if not construction_var.get():
            project_types.remove('Construction Administration')
        if not other_var.get():
            project_types.remove('Other')
        project_types.remove('None')#if there is any spec, remove None.
        return project_types    

def get_location():
    '''
    This function returns all the location specific information
    '''
    radius = int(loc_radius_var.get())
    start = str(loc_entry_var.get())
    print(start)
    if location_var_1.get() == "By Address":
        time.sleep(1)#Respect for Nominatim
        location = geolocator.geocode(start)
        coords = (location.latitude,location.longitude)
        return coords
    else:
        #print("BY ID")
        database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)
        database_sheet = database_wb[DATABASE_SHEET_NAME]
        center_lat = 0
        center_long = 0
        for x in range(2,database_sheet.max_row+1):
            if database_sheet["A"+str(x)].value == start:
                coords = (database_sheet["G"+str(x)].value,database_sheet["H"+str(x)].value)
                return coords
    #I could maybe add these to a list and use that? 
        
    
def excel_gen():
    '''
    When EXCEL button is pressed
    '''
    years = get_years() #list of years to select
    clients = get_client_types() #list of client types
    projects = get_project_types()
    if loc_spec_var.get():
        #if it is get the location to select by
        coords = get_location()
        radius = int(loc_radius_var.get())
    database_wb = openpyxl.load_workbook(DATABASE_SPREADSHEET)
    database_sheet = database_wb[DATABASE_SHEET_NAME]
    
    write_sheet_row = 2
    database_wb.create_sheet(index=0,title='Output')
    write_sheet = database_wb['Output']
    write_sheet["A1"].value="Project ID"
    write_sheet["B1"].value="Project Name"
    write_sheet["C1"].value="Location"
    write_sheet["D1"].value="Project Type"
    write_sheet["E1"].value="Client"
    write_sheet["F1"].value="Client Type"
    write_sheet["G1"].value="Lat"
    write_sheet["H1"].value="Long"
    write_sheet["I1"].value="Permit"
    write_sheet["J1"].value="Boring"
    write_sheet["K1"].value="Distance (mi)"

    for i in range(2,database_sheet.max_row):
        proj_id = database_sheet["A"+str(i)].value
        proj_year = str(proj_id)[:4]
        if int(proj_year) in years:#check if meets year criteria
            if database_sheet["D"+str(i)].value in projects:#check if meets project type criteria
                if database_sheet["F"+str(i)].value in clients:#check if meets client type criteria
                    if not loc_spec_var.get():#if we do NOT need to sort by location
                        #copy to excel sheet
                        write_sheet["A"+str(write_sheet_row)].value = database_sheet["A"+str(i)].value
                        write_sheet["B"+str(write_sheet_row)].value = database_sheet["B"+str(i)].value
                        write_sheet["C"+str(write_sheet_row)].value = database_sheet["C"+str(i)].value
                        write_sheet["D"+str(write_sheet_row)].value = database_sheet["D"+str(i)].value
                        write_sheet["E"+str(write_sheet_row)].value = database_sheet["E"+str(i)].value
                        write_sheet["F"+str(write_sheet_row)].value = database_sheet["F"+str(i)].value
                        write_sheet["G"+str(write_sheet_row)].value = database_sheet["G"+str(i)].value
                        write_sheet["H"+str(write_sheet_row)].value = database_sheet["H"+str(i)].value
                        write_sheet_row +=1
                    else:
                        #test to see if it is within radius
                        #distance = geodesic(tuple1lat,long,tuple2lat,long).miles
                        if database_sheet["G"+str(i)].value != None:
                            t_coords = (float(database_sheet["G"+str(i)].value),float(database_sheet["H"+str(i)].value))
                            distance = geodesic(coords,t_coords).miles
                            print(distance)
                            if distance < radius:
                                write_sheet["A"+str(write_sheet_row)].value = database_sheet["A"+str(i)].value
                                write_sheet["B"+str(write_sheet_row)].value = database_sheet["B"+str(i)].value
                                write_sheet["C"+str(write_sheet_row)].value = database_sheet["C"+str(i)].value
                                write_sheet["D"+str(write_sheet_row)].value = database_sheet["D"+str(i)].value
                                write_sheet["E"+str(write_sheet_row)].value = database_sheet["E"+str(i)].value
                                write_sheet["F"+str(write_sheet_row)].value = database_sheet["F"+str(i)].value
                                write_sheet["G"+str(write_sheet_row)].value = database_sheet["G"+str(i)].value
                                write_sheet["H"+str(write_sheet_row)].value = database_sheet["H"+str(i)].value
                                write_sheet["k"+str(write_sheet_row)].value = distance
                                write_sheet_row +=1
    database_wb.remove(database_wb[DATABASE_SHEET_NAME])
    database_wb.save('RACEGIS_OUTPUT.xlsx')
    os.startfile("RACEGIS_OUTPUT.xlsx")


def all_client():
    if all_client_var.get():
        residential_var.set(1)
        government_var.set(1)
        myc_var.set(1) 
        util_var.set(1) 
        consultant_var.set(1)
        contractor_var.set(1)
        print("Selecting all clients")
    else:
        residential_var.set(0)
        government_var.set(0)
        myc_var.set(0) 
        util_var.set(0) 
        consultant_var.set(0)
        contractor_var.set(0)
        print("deselecting all clients")

def client():
    if residential_var.get() and \
         government_var.get() and \
         myc_var.get() and \
         util_var.get() and \
         consultant_var.get() and \
         contractor_var.get():
            all_client_var.set(1)
    else:
        all_client_var.set(0)


def all_proj():
    if all_proj_var.get():
        dredge_var.set(1)
        floodwave_var.set(1)
        hydro_var.set(1) 
        pierdock_var.set(1) 
        floodcontrol_var.set(1)
        upland_var.set(1)
        facilities_var.set(1)
        beach_var.set(1)
        construction_var.set(1)
        other_var.set(1)
        print("Selecting all projects")

    else:
        dredge_var.set(0)
        floodwave_var.set(0)
        hydro_var.set(0) 
        pierdock_var.set(0) 
        floodcontrol_var.set(0)
        upland_var.set(0)
        facilities_var.set(0)
        beach_var.set(0)
        construction_var.set(0)
        other_var.set(0)
        print("deselecting all projects")

    
def proj():
    '''
    Dummy function, the states will be checked later, all this does is set all on or off
    It checks if all the projects are checked, puts all on
    if all are not checked, leaves (or sets) it unchecked.
    '''
    if dredge_var.get() and \
         floodwave_var.get() and \
         hydro_var.get() and \
         pierdock_var.get() and \
         floodcontrol_var.get() and \
         upland_var.get() and \
         facilities_var.get() and \
         beach_var.get() and \
         construction_var.get() and \
         other_var.get():
            all_proj_var.set(1)
    else:
        all_proj_var.set(0)

def location_spec():
    '''
    This function checks the status of the 'specify location' check box
    is the box is checked (.get() returns true) then the state of all the location boxes is set to normal
    meaning that they are no longer greyed out.
    If the box is unchecked, they will all revert to being greyed out.
    '''
    if loc_spec_var.get():
        location_drop_down.config(state=tk.NORMAL)
        loc_entry.config(state=tk.NORMAL)
        loc_radius.config(state=tk.NORMAL)
        loc_radius_label.config(state=tk.NORMAL)
        starting_point_label.config(state=tk.NORMAL)


    else:
        location_drop_down.config(state=tk.DISABLED)    
        loc_entry.config(state=tk.DISABLED)
        loc_radius.config(state=tk.DISABLED)    
        loc_radius_label.config(state=tk.DISABLED)
        starting_point_label.config(state=tk.DISABLED)


      
def checkyear(year):
    if year_var_2.get() == "1996":
        year_var_1.set(year_var_2.get())
    elif year_var_2.get() < year_var_1.get():
        year2=year_var_2.get()
        year2=int(year2)
        year_var_1.set(year2-1)


    #Check to see if year_var_1 > year_var 2, change year_var 1 to be a year less, unless the year is 1996
    


#Create the window    
root = tk.Tk()
#Set the titlebar
root.title("RACE GIS v"+str(VERSION))

#YEARS
year_choices = []
year_var_1 = tk.StringVar()
year_var_1.set(str(now.year))
year_var_2 = tk.StringVar()
year_var_2.set(str(now.year))
for x in range(1996,now.year+1):
    year_choices.append(x)
    
years = ttk.Labelframe(root,text="Years")
to_label = tk.Label(years, text="to")
from_label = tk.Label(years, text="From")


years.grid(row=0, column=1,columnspan=2,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)
from_label.grid(row=0,column=1)
year_menu_1 = tk.OptionMenu(years,year_var_1, *year_choices,command=checkyear)
year_menu_1.grid(row=0,column=2)
to_label.grid(row=0,column=3)

year_menu_2 = tk.OptionMenu(years,year_var_2, *year_choices,command=checkyear)
year_menu_2.grid(row=0,column=4)


#PROJECTS
projects = ttk.Labelframe(root,text="Project Types")
projects.grid(row=1, column=1,sticky=tk.N,ipadx=2,ipady=2,padx=2,pady=2)

all_proj_var = tk.IntVar(value=1)
all_proj_checkbox = tk.Checkbutton(projects, text="All",variable=all_proj_var,command=all_proj)
all_proj_checkbox.grid(row=0, column=1,sticky=tk.W)

dredge_var = tk.IntVar(value=1)
dredge_checkbox = tk.Checkbutton(projects, text="Dredging",variable=dredge_var,command=proj)
dredge_checkbox.grid(row=1, column=1,sticky=tk.W)

floodwave_var = tk.IntVar(value=1)
floodwave_checkbox = tk.Checkbutton(projects, text="Flood / Wave Study",variable=floodwave_var,command=proj)
floodwave_checkbox.grid(row=2, column=1,sticky=tk.W)

hydro_var = tk.IntVar(value=1)
hydro_checkbox = tk.Checkbutton(projects, text="Hydrographic Survey",variable=hydro_var,command=proj)
hydro_checkbox.grid(row=3, column=1,sticky=tk.W)

pierdock_var = tk.IntVar(value=1)
pierdock_checkbox = tk.Checkbutton(projects, text="Pier / Dock",variable=pierdock_var,command=proj)
pierdock_checkbox.grid(row=4, column=1,sticky=tk.W)

floodcontrol_var = tk.IntVar(value=1)
floodcontrol_checkbox = tk.Checkbutton(projects, text="Flood / Erosion Control",variable=floodcontrol_var,command=proj)
floodcontrol_checkbox.grid(row=5, column=1,sticky=tk.W)

upland_var = tk.IntVar(value=1)
upland_checkbox = tk.Checkbutton(projects, text="Upland Structure",variable=upland_var,command=proj)
upland_checkbox.grid(row=6, column=1,sticky=tk.W)

facilities_var = tk.IntVar(value=1)
facilities_checkbox = tk.Checkbutton(projects, text="Facilities Assessment",variable=facilities_var,command=proj)
facilities_checkbox.grid(row=7, column=1,sticky=tk.W)

beach_var = tk.IntVar(value=1)
beach_checkbox = tk.Checkbutton(projects, text="Beach Nourishment",variable=beach_var,command=proj)
beach_checkbox.grid(row=8, column=1,sticky=tk.W)

construction_var = tk.IntVar(value=1)
construction_checkbox = tk.Checkbutton(projects, text="Construction Admin",variable=construction_var,command=proj)
construction_checkbox.grid(row=9, column=1,sticky=tk.W)

other_var = tk.IntVar(value=1)
other_checkbox = tk.Checkbutton(projects, text="Other",variable=other_var,command=proj)
other_checkbox.grid(row=10, column=1,sticky=tk.W)
#############

###CLIENTS###
clients = ttk.Labelframe(root,text="Client Types")
clients.grid(row=1,column=2,sticky=tk.N+tk.S,ipadx=2,ipady=2,padx=2,pady=2)

all_client_var = tk.IntVar(value=1)
all_proj_checkbox = tk.Checkbutton(clients, text="All",variable=all_client_var,command=all_client)
all_proj_checkbox.grid(row=0, column=1,sticky=tk.W)

residential_var = tk.IntVar(value=1)
residential_checkbox = tk.Checkbutton(clients, text="Residential",variable=residential_var,command=client)
residential_checkbox.grid(row=1, column=1,sticky=tk.W)

government_var = tk.IntVar(value=1)
government_checkbox = tk.Checkbutton(clients, text="Goverment",variable=government_var,command=client)
government_checkbox.grid(row=2, column=1,sticky=tk.W)

myc_var = tk.IntVar(value=1)
myc_checkbox = tk.Checkbutton(clients, text="Marina / Yacht Club",variable=myc_var,command=client)
myc_checkbox.grid(row=3, column=1,sticky=tk.W)

util_var = tk.IntVar(value=1)
util_checkbox = tk.Checkbutton(clients, text="Industrial / Utility",variable=util_var,command=client)
util_checkbox.grid(row=4, column=1,sticky=tk.W)

consultant_var = tk.IntVar(value=1)
consultant_checkbox = tk.Checkbutton(clients, text="Consultant",variable=consultant_var,command=client)
consultant_checkbox.grid(row=5, column=1,sticky=tk.W)

contractor_var = tk.IntVar(value=1)
contractor_checkbox = tk.Checkbutton(clients, text="Contractor",variable=contractor_var,command=client)
contractor_checkbox.grid(row=6, column=1,sticky=tk.W)
#############

#EXTRAS
extras = ttk.Labelframe(root,text="Extras")
extras.grid(row=2,column=2,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)

permit_var = tk.IntVar(value=0)
permit_checkbox = tk.Checkbutton(extras, text="Permit",variable=permit_var)
permit_checkbox.grid(row=0, column=1,sticky=tk.W)
permit_checkbox.config(state=tk.DISABLED)   #remove once added 

boring_var = tk.IntVar(value=0)
boring_checkbox = tk.Checkbutton(extras, text="Boring",variable=boring_var)
boring_checkbox.grid(row=1, column=1,sticky=tk.W)
boring_checkbox.config(state=tk.DISABLED)  #remove once added  

###DEEP ZONES###
deep = ttk.Labelframe(root,text="CT DEEP Zones")
deep.grid(row=2,sticky=tk.W+tk.E, column=1,ipadx=2,ipady=2,padx=2,pady=2)

aquifer_var = tk.IntVar(value=0)
aquifer_checkbox = tk.Checkbutton(deep, text="Aquifer Protection",variable=aquifer_var)
aquifer_checkbox.grid(row=0, column=1,sticky=tk.W)
aquifer_checkbox.config(state=tk.DISABLED)  #remove once added  

diversity_var = tk.IntVar(value=0)
diversity_checkbox = tk.Checkbutton(deep, text="Diversity Database",variable=diversity_var)
diversity_checkbox.grid(row=1, column=1,sticky=tk.W)
diversity_checkbox.config(state=tk.DISABLED)  #remove once added  

#############


##VISUALS##
visuals = ttk.Labelframe(root, text="Visuals")
visuals.grid(row=3,column=1,sticky=tk.W+tk.E,columnspan=2,ipadx=2,ipady=2,padx=2,pady=2)
heatmap_checkbox_var = tk.IntVar(value=0)
heatmap_checkbox = tk.Checkbutton(visuals, text="Heatmap",variable=heatmap_checkbox_var)
heatmap_checkbox.grid(row=0, column=1,sticky=tk.W)
heatmap_checkbox.config(state=tk.DISABLED)


#LOCATION
location_by_options = ['By Address','By ID']
location = ttk.Labelframe(root,text="Location")
location.grid(row=4,column=1,sticky=tk.W+tk.E,columnspan=2,ipadx=2,ipady=2,padx=2,pady=2)

loc_spec_var = tk.IntVar(value=0)
location_spec = tk.Checkbutton(location,text="Specify Location", command=location_spec, var=loc_spec_var)
location_spec.grid(row=0,column=1)

location_var_1 = tk.StringVar()
location_var_1.set('By ID')
location_drop_down = tk.OptionMenu(location,location_var_1, *location_by_options)

location_drop_down.grid(column=2,row=0,sticky=tk.W+tk.E)
location_drop_down.config(state=tk.DISABLED)

starting_point_label = tk.Label(location,text="Starting Point")
starting_point_label.grid(column=1,row=2,sticky=tk.W+tk.E)
starting_point_label.config(state=tk.DISABLED)

loc_entry_var = tk.StringVar()
loc_entry = tk.Entry(location,textvariable=loc_entry_var)
loc_entry.config(state=tk.DISABLED)
loc_entry.config(font=("Arial",10))

loc_entry.grid(column=2,row=2,sticky=tk.W+tk.E)

loc_radius_label = tk.Label(location,text="Radius (mi)")
loc_radius_label.grid(column=1,row=3,sticky=tk.W+tk.E)
loc_radius_label.config(state=tk.DISABLED)

loc_radius_var = tk.StringVar()
loc_radius = tk.Entry(location,textvariable=loc_radius_var)
loc_radius.config(state=tk.DISABLED)    
loc_radius_var.set("10")
loc_radius.grid(column=2,row=3)

#MAP
map_button = tk.Button(root,text='Google Earth',command=map_gen)
map_button.grid(row=5,column=1,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)

#SPREADSHEET
excel_button = tk.Button(root,text='Excel',command=excel_gen)
excel_button.grid(row=5,column=2,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)

if check_for_spreadsheets():
    check_database()

root.mainloop()
