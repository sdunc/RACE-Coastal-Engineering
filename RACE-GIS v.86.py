#Stephen Duncanon
#RACE GIS 
#7/12/19 -

import time
import os
import os.path
import tkinter as tk
from tkinter import ttk
import datetime
import time
import openpyxl
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
                     
DATABASE_HEADER =  ['Project ID',       #0 / A
                    'Project Name',     #1 / B
                    'Location',         #2 / C
                    'Project Type',     #3 / D
                    'Client',           #4 / E
                    'Client Type',      #5 / F
                    'Latitude',         #6 / G
                    'longitude']        #7 / H
                    #'Permit',          #8 adding this here for
                    #'Boring']          #9 future proofing purposes!


PROJECTS_FILE       = 'X:/RACEGIS/projects.xlsx'    #The projects source file location
CLIENTS_FILE        = 'X:/RACEGIS/clients.xlsx'     #The clients source file location
DATABASE_FILE       = 'X:/RACEGIS/database.xlsx'    #The location of the database file
CLIENTS_SHEET_NAME  = 'GIS_Clients'                 #this is equal to the name set by the custom inquiry in AJERA
PROJECTS_SHEET_NAME = 'GIS_Projects'                #this is equal to the name set by the custom inquiry in AJERA
DATABASE_SHEET_NAME = 'data'                        #name of the sheet inside database
VERSION = .86                                       #Version Number

#Variables
now = datetime.datetime.now()
kml_header_1='<?xml version="1.0" encoding="UTF-8"?>'
kml_header =\
'''<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>RACE GIS</name>
  <Style id="project-style">
  <LabelStyle>
   <scale>0</scale>
  </LabelStyle>
   <IconStyle>
    <scale>1</scale>
    <Icon>
      <href>http://maps.google.com/mapfiles/kml/paddle/red-circle.png</href>
    </Icon>
    </IconStyle>
    <BalloonStyle>
      <text>
        <![CDATA[
	  <b>$[name]</b>
      <hr>
          <p><b>ID: </b>$[ID]</p>
	  <p><b>Address: </b>$[Address]</p>
          <p><b>Project Type: </b>$[projectType]</p>
          <p><b>Client Name: </b>$[clientName]</p>
          <p><b>Client Type: </b>$[clientType]</p>
        ]]>
      </text>
    </BalloonStyle>
  </Style>
'''
kml_footer=\
'''</Document>
</kml>'''

#Functions
def startup_database_check():
    '''
    This function is called every time that the project is opened.
    It handles the fixing/updating of the database
    '''
    if is_database_present():
        if not is_database_updated():
            update_database("update")
    else:
        if are_sources_present():
            create_database()
        else:
            #If the sources AND database are missing, statusbar_text.set error to statusbar, and disable buttons
            popup("Error! Sources are missing")
            


def is_database_present():
    '''
    This function determines if the database exists or not
    '''
    if os.path.exists(DATABASE_FILE):
        statusbar_text.set("database found!")
        
        return True
    else:
        return False

def is_database_updated():
    '''
    This function checks if the database is updated by counting the lines in the database
    This is easier to do now that we are returning to a spreadsheet based model since I can use the max_row method to quickly get the max rows
    '''
    #Start by opening up both files, this code creates two objects 
    projects_file = openpyxl.load_workbook(PROJECTS_FILE)
    database = openpyxl.load_workbook(DATABASE_FILE)
    #Open the specific sheets that I want to examine
    projects_sheet = projects_file[PROJECTS_SHEET_NAME]
    database_sheet = database[DATABASE_SHEET_NAME]   
    
    #Get the number of rows in each database
    rows_in_projects_sheet = int(projects_sheet.max_row)
    rows_in_database_sheet = int(database_sheet.max_row)
    
    #Now there are 3 cases, if the database rows = project rows, updated
    #if projects > database, it needs to update
    #otherwise error! 
    
    if rows_in_projects_sheet == rows_in_database_sheet:
        statusbar_text.set("The database is up to date!")    #this can be removed
        
        return True
    elif rows_in_projects_sheet > rows_in_database_sheet:
        statusbar_text.set("FALSE. Database needs to be updated") #this can be removed
        
        return False
    else:
       statusbar_text.set("Error, database is ahead of project source, outdated project files?")
       

def are_sources_present():
    if os.path.exists(CLIENTS_FILE) and os.path.exists(PROJECTS_FILE):
        return True
    else:
        return False

def create_database():
    '''
    The create database function is called when no database is found but source files are present
    it creates a new text file and writes the header constant to it before calling the update database function
    '''
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0,title=DATABASE_SHEET_NAME)
    database_sheet = wb[DATABASE_SHEET_NAME]
    wb.remove(wb['Sheet'])
    
    #The list of columns to iterate over, there is probably a better way to do this, maybe something built into Openpyxl?
    column_letter = ["A","B","C","D","E","F","G","H","I","J","K","L"]
    
    #for each item in constant list, add it to the first entry of the column letter list, then remove that letter so the next
    #item is added to the next column.
    for header_item in DATABASE_HEADER:
        database_sheet[column_letter[0]+"1"].value = header_item
        column_letter.pop(0)

    wb.save(DATABASE_FILE)
    statusbar_text.set('Database created')
    
    update_database("new")
    
def update_database(new_or_update):
    add_new_projects()
    add_client_types()
    geocode_database()
    #I added a flag here to avoid the possible duplication issue that might happen if the database attempts to fix itself after being generated
    if new_or_update == "update":
        fix_database()
   

def add_new_projects():
    '''
    The add_new_projects function is called when the database is being updated
    I am using the same algorithm I used in v0.8, a triple pass
    1) add all Ids from the project file to a list
    2) remove all Ids that are also in the first column of the database
    3) go through the project file and find the rows with those ids that remain in the list, adding their contents to the database
    this could probably be sped up in a number of ways
    copying a range? grabbing the whole row? finding a way to iterate fewer times?
    actually I just thought of one
    iterate over the database sheet and add all the current ids to a list
    go into the project files and if the id is not in that list, copy it to the current (end) line of the databse file! 
    much cleaner ALSO check to see if the database is empty (was just created!) if it was, skip right to adding everything
    maybe this could be done faster still by doing it in the create_database() function and maybe just changing the header of the project file
    '''
    
    #One thing I might want to do here is to make the projects_file readonly to increase speed and avoid the possibility that it gets messed up
    projects_file = openpyxl.load_workbook(PROJECTS_FILE)
    database = openpyxl.load_workbook(DATABASE_FILE)
    #Open the specific sheets that I want to examine
    projects_sheet = projects_file[PROJECTS_SHEET_NAME]
    database_sheet = database[DATABASE_SHEET_NAME]  
    
    
    current_database_write_row = database_sheet.max_row+1 #go to the next row after the final row of the database
    
    if database_sheet.max_row ==1: #if the database was just created and just has a header
        #write everything without needing to check what goes in
        for i in range(2,projects_sheet.max_row+1):                                                       #All the things that will be in the projects export
            database_sheet["A"+str(current_database_write_row)].value=projects_sheet["A"+str(i)].value   #id
            database_sheet["B"+str(current_database_write_row)].value=projects_sheet["B"+str(i)].value   #Name
            database_sheet["C"+str(current_database_write_row)].value=projects_sheet["C"+str(i)].value    #Location
            database_sheet["D"+str(current_database_write_row)].value=projects_sheet["D"+str(i)].value   #Project Type
            database_sheet["E"+str(current_database_write_row)].value=projects_sheet["E"+str(i)].value   #Client
            current_database_write_row +=1 #move to the next row 
    else:
        current_database_ids = []
        #if the database already has some things in it and we need to check what needs to be added
        for j in range(2,database_sheet.max_row+1):
            #add all the current ids to a list
            current_database_ids.append(database_sheet["A"+str(j)].value)
            
        for k in range(2,projects_sheet.max_row+1):
            if str(projects_sheet["A"+str(k)].value) not in str(current_database_ids):
                database_sheet["A"+str(current_database_write_row)].value=projects_sheet["A"+str(k)].value
                database_sheet["B"+str(current_database_write_row)].value=projects_sheet["B"+str(k)].value
                database_sheet["C"+str(current_database_write_row)].value=projects_sheet["C"+str(k)].value
                database_sheet["D"+str(current_database_write_row)].value=projects_sheet["D"+str(k)].value
                database_sheet["E"+str(current_database_write_row)].value=projects_sheet["E"+str(k)].value
                current_database_write_row +=1 #move to the next row 

                #PERMITS AND BORING
                #database_sheet["G"+str(current_database_row)].value=projects_sheet["G"+str(k)].value
                #database_sheet["H"+str(current_database_row)].value=projects_sheet["H"+str(k)].value
    statusbar_text.set('New projects added to database')
    
    database.save(DATABASE_FILE)
   
def add_client_types():
    '''
    create a dictionary with key value pairs
    then iterate through database, breaking each line into a list
    search each list for the key, if found, add the value into the nested list
    write each list back to the file delimited
    '''
    clients = openpyxl.load_workbook(CLIENTS_FILE)
    database = openpyxl.load_workbook(DATABASE_FILE)
    client_sheet = clients[CLIENTS_SHEET_NAME]
    database_sheet = database[DATABASE_SHEET_NAME]
    
    d = {}

    for x in range(2,client_sheet.max_row+1):
        d[client_sheet["A"+str(x)].value] = str(client_sheet["B"+str(x)].value)
        
    for i in range(2,database_sheet.max_row+1):
        database_sheet["F"+str(i)].value= d[database_sheet["E"+str(i)].value]   
     
    statusbar_text.set('Client types added to database')
    
    database.save(DATABASE_FILE)
  
def geocode_database():
    '''
    This function will iterate through the list of jobs (nested list)
    and find ones with a len() too short to include lat/long
    it will grab the location field and use nominatim to geocode
    using try/except (TypeError) it will either put the address or (0,0)
    '''
    log_file = open('X:\RACEGIS\logs\log.txt','a')
    statusbar_text.set('Staring bulk geocode, This will take several minutes!')
    
    database = openpyxl.load_workbook(DATABASE_FILE)
    database_sheet = database[DATABASE_SHEET_NAME]
    
    geolocator = Nominatim(timeout=10,user_agent="RACE-GIS")

    
    for x in range(2,database_sheet.max_row+1):
        if database_sheet["G"+str(x)].value == None: #meaning that it has no value for lat
            time.sleep(1)#respect for Nominatim
            
            try:
                location = geolocator.geocode(str(database_sheet["C"+str(x)].value)) #attempt to geocode the location cell for current row
                latitude  = location.latitude
                longitude = location.longitude
                database_sheet["G"+str(x)].value = latitude
                database_sheet["H"+str(x)].value = longitude
                statusbar_text.set(str(database_sheet["A"+str(x)].value)+" was geocoded!")
            except:
                AttributeError
                database_sheet["G"+str(x)].value = 0 #if the address could not be geocoded, write 0,0 as lat and long
                database_sheet["H"+str(x)].value = 0
                log_file.write(str(database_sheet["A"+str(x)].value)+" could not be geocoded!") #write the ids of the projects that could not be geocoded
                statusbar_text.set(str(database_sheet["A"+str(x)].value)+" could not be geocoded!")

    database.save(DATABASE_FILE)
               
def fix_database():
    '''
    This function will run after the database has been updated OR
    when the user chooses this option from the menu,
    that would be if the export was updated but no new fields had been added
    it will search through the database and look for any projects with
    0 for lat and 0 for long, (placeholders) then it will attempt to geocode the address
    field and overwrite with true values
    '''
    log_file = open('X:\RACEGIS\logs\log.txt','a')
    
    database = openpyxl.load_workbook(DATABASE_FILE)
    database_sheet = database[DATABASE_SHEET_NAME]
    
    geolocator = Nominatim(timeout=10,user_agent="RACE-GIS")   
 
    for x in range(2,database_sheet.max_row):
        if str(database_sheet["G"+str(x)].value) == "0":
            time.sleep(1)#respect for Nominatim
            try:
                location = geolocator.geocode(str(database_sheet["C"+str(x)].value)) #attempt to geocode the location cell for current row
                latitude  = location.latitude
                longitude = location.longitude
                database_sheet["G"+str(x)].value = latitude
                database_sheet["H"+str(x)].value = longitude
                statusbar_text.set(str(database_sheet["A"+str(x)].value)+" was fixed!")
            except:
                AttributeError
                database_sheet["G"+str(x)].value = 0 #if the address could not be geocoded, write 0,0 as lat and long
                database_sheet["H"+str(x)].value = 0
                statusbar_text.set(str(database_sheet["A"+str(x)].value)+" broken geocode could not be fixed!")
                log_file.write(str(database_sheet["A"+str(x)].value)+" broken geocode could not be fixed!") #write the ids of the projects that could not be geocoded 
                
    database.save(DATABASE_FILE)
    

def google_earth_button():
    '''
    When map button is pressed
    '''
    database_wb = openpyxl.load_workbook(DATABASE_FILE)
    database_sheet = database_wb[DATABASE_SHEET_NAME]

    projects = get_project_types()
    years = get_years() #list of years to select
    clients = get_client_types() #list of client types

    #check if location is checked
    if loc_spec_var.get():
        #if it is get the location to select by
        coords = get_location()
        radius = int(loc_radius_var.get())
    export_file = open('RACEGIS.kml','w')#open a text file of the date + rest
    export_file.write(kml_header_1)
    export_file.write(kml_header)
    for i in range(2,database_sheet.max_row):
        #FILTER
        proj_id = database_sheet["A"+str(i)].value
        proj_year = str(proj_id)[:4]
        if int(proj_year) in years:#check if meets year criteria
            if database_sheet["D"+str(i)].value in projects:#check if meets project type criteria
                if database_sheet["F"+str(i)].value in clients:#check if meets client type criteria
                    if not loc_spec_var.get():#if we do NOT need to sort by location
                                
                        name         =  str(database_sheet["B"+str(i)].value)
                        if "&" in name:
                            name = name.replace("&","/")
                            
                        proj_id      =  str(database_sheet["A"+str(i)].value)
                        if "&" in proj_id:
                            proj_id = proj_id.replace("&","/")
                            
                        location     =  str(database_sheet["C"+str(i)].value)
                        if "&" in location:
                            location     =  location.replace("&","/")

                        proj_type    =  str(database_sheet["D"+str(i)].value)
                        if "&" in proj_type:
                            proj_type = proj_type.replace("&","/")

                        client       =  str(database_sheet["E"+str(i)].value)
                        if "&" in client:
                            client = client.replace("&","/")
                        
                        client_type  =  str(database_sheet["F"+str(i)].value)
                        if "&" in client_type:
                            client_type = client_type.replace("&","/")
                                
                        long_lat_str =  str(database_sheet["H"+str(i)].value)+","+str(database_sheet["G"+str(i)].value)
                        export_file.write('  <Placemark>\n') 
                        export_file.write('    <name>'+name+'</name>\n')
                        export_file.write('    <styleUrl>#project-style</styleUrl>\n')
                        export_file.write('    <ExtendedData>\n')
                        export_file.write('      <Data name="ID">\n')
                        export_file.write('	<value>'+proj_id+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('      <Data name="Address">\n')
                        export_file.write('        <value>'+location+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('      <Data name="projectType">\n')
                        export_file.write('        <value>'+proj_type+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('      <Data name="clientType">\n')
                        export_file.write('        <value>'+client_type+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('      <Data name="clientName">\n')
                        export_file.write('        <value>'+client+'</value>\n')
                        export_file.write('      </Data>\n')
                        export_file.write('    </ExtendedData>\n')
                        export_file.write('    <Point>\n')
                        export_file.write('         <coordinates>'+long_lat_str+',0</coordinates>\n')
                        export_file.write('    </Point>\n')
                        export_file.write('  </Placemark>\n')

                    else:
                        if database_sheet["G"+str(i)].value != None:
                            t_coords = (float(database_sheet["G"+str(i)].value),float(database_sheet["H"+str(i)].value))
                            distance = geodesic(coords,t_coords).miles
                            statusbar_text.set(distance)
                            if distance < radius:
                                
                                name         =  str(database_sheet["B"+str(i)].value)
                                if "&" in name:
                                    name = name.replace("&","/")
                                    
                                proj_id      =  str(database_sheet["A"+str(i)].value)
                                if "&" in proj_id:
                                    proj_id = proj_id.replace("&","/")
                                    
                                location     =  str(database_sheet["C"+str(i)].value)
                                if "&" in location:
                                    location     =  location.replace("&","/")

                                proj_type    =  str(database_sheet["D"+str(i)].value)
                                if "&" in proj_type:
                                    proj_type = proj_type.replace("&","/")
 
                                client       =  str(database_sheet["E"+str(i)].value)
                                if "&" in client:
                                    client = client.replace("&","/")
                                
                                client_type  =  str(database_sheet["F"+str(i)].value)
                                if "&" in client_type:
                                    client_type = client_type.replace("&","/")
                                
                                long_lat_str =  str(database_sheet["H"+str(i)].value)+","+str(database_sheet["G"+str(i)].value)
                                export_file.write('  <Placemark>\n') 
                                export_file.write('    <name>'+name+'</name>\n')
                                export_file.write('    <styleUrl>#project-style</styleUrl>\n')
                                export_file.write('    <ExtendedData>\n')
                                export_file.write('      <Data name="ID">\n')
                                export_file.write('	<value>'+proj_id+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="Address">\n')
                                export_file.write('        <value>'+location+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="projectType">\n')
                                export_file.write('        <value>'+proj_type+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="clientType">\n')
                                export_file.write('        <value>'+client_type+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('      <Data name="clientName">\n')
                                export_file.write('        <value>'+client+'</value>\n')
                                export_file.write('      </Data>\n')
                                export_file.write('    </ExtendedData>\n')
                                export_file.write('    <Point>\n')
                                export_file.write('         <coordinates>'+long_lat_str+',0</coordinates>\n')
                                export_file.write('    </Point>\n')
                                export_file.write('  </Placemark>\n')
    export_file.write(kml_footer)
    export_file.close()
    os.startfile("RACEGIS.kml")


def get_years():
    years = []
    for x in range(int(year_var_1.get()),int(year_var_2.get())+1):
        years.append(x)
    return years

def get_client_types():
    client_types = ['Residential',
                    'Government',
                    'Marina / Yacht Club',
                    'Industrial / Utility',
                    'Consultant',
                    'Contractor']
    if all_client_var.get():
        return client_types
    elif not residential_var.get() and \
         not government_var.get() and \
         not myc_var.get() and \
         not util_var.get() and \
         not consultant_var.get() and \
         not contractor_var.get():
            statusbar_text.set("ERROR! SELECT a client type")
    else:
        if not residential_var.get():
            client_types.remove('Residential')
        if not government_var.get():
            client_types.remove('Government')
        if not myc_var.get():
            client_types.remove('Marina / Yacht Club')
        if not util_var.get():
            client_types.remove('Industrial / Utility')
        if not consultant_var.get():
            client_types.remove('Consultant')
        if not contractor_var.get():
            client_types.remove('Contractor')
        return client_types

def get_project_types():
    project_types = ['Dredging',#0
                     'Flood and erosion control structure',#1
                     'Beach / Dune / Living Shoreline',#2
                     'Pier / Dock',#3
                     'Hydrographic Survey',#4
                     'Facility Assessment',#5
                     'Construction Administration',#6
                     'Flood / Wave Study',#7
                     'Upland Structure',#8
                     'Other',#9
                     'None'#Source of error
                     ]
    #check each checkbox, if not checked, remove from project types list
    #statusbar_text.set list of projects that are wanted
    #first check if all is cheched
    if all_proj_var.get():
        return project_types
        #this case quickly sees if all projects are checked
        #since if all is checked they all must be?
    elif not dredge_var.get() and \
         not floodwave_var.get() and \
         not hydro_var.get() and \
         not pierdock_var.get() and \
         not floodcontrol_var.get() and \
         not upland_var.get() and \
         not facilities_var.get() and \
         not beach_var.get() and \
         not construction_var.get() and \
         not other_var.get():
            statusbar_text.set("ERROR! SELECT a project type")
    else:
        if not dredge_var.get():
            project_types.remove('Dredging')
        if not floodwave_var.get():
            project_types.remove('Flood / Wave Study')
        if not hydro_var.get():
            project_types.remove('Hydrographic Survey')
        if not pierdock_var.get():
            project_types.remove('Pier / Dock')
        if not floodcontrol_var.get():
            project_types.remove('Flood and erosion control structure')
        if not upland_var.get():
            project_types.remove('Upland Structure')
        if not facilities_var.get():
            project_types.remove('Facility Assessment')
        if not beach_var.get():
            project_types.remove('Beach / Dune / Living Shoreline')
        if not construction_var.get():
            project_types.remove('Construction Administration')
        if not other_var.get():
            project_types.remove('Other')
        return project_types    

def get_location():
    '''
    This function returns all the location specific information
    '''
    radius = int(loc_radius_var.get())
    start = str(loc_entry_var.get())
    statusbar_text.set(start)
    if location_var_1.get() == "By Address":
        geolocator = Nominatim(timeout=10,user_agent="RACE-GIS")
        location = geolocator.geocode(start)
        coords = (location.latitude,location.longitude)
        return coords
    else:
        #statusbar_text.set("BY ID")
        database_wb = openpyxl.load_workbook(DATABASE_FILE)
        database_sheet = database_wb[DATABASE_SHEET_NAME]
        for x in range(2,database_sheet.max_row+1):
            if str(database_sheet["A"+str(x)].value) == start:
                coords = (database_sheet["G"+str(x)].value,database_sheet["H"+str(x)].value)
                return coords
    #I could maybe add these to a list and use that? 
        
    
def excel_button():
    '''
    When EXCEL button is pressed
    '''
    years = get_years() #list of years to select
    clients = get_client_types() #list of client types
    projects = get_project_types()
    if loc_spec_var.get():
        #if it is get the location to select by
        coords = get_location()
        radius = int(loc_radius_var.get())
    database_wb = openpyxl.load_workbook(DATABASE_FILE)
    database_sheet = database_wb[DATABASE_SHEET_NAME]
    
    write_sheet_row = 2
    database_wb.create_sheet(index=0,title='Output')
    write_sheet = database_wb['Output']
    write_sheet["A1"].value="Project ID"
    write_sheet["B1"].value="Project Name"
    write_sheet["C1"].value="Location"
    write_sheet["D1"].value="Project Type"
    write_sheet["E1"].value="Client"
    write_sheet["F1"].value="Client Type"
    write_sheet["G1"].value="Lat"
    write_sheet["H1"].value="Long"
    write_sheet["I1"].value="Permit"
    write_sheet["J1"].value="Boring"
    write_sheet["K1"].value="Distance (mi)"

    for i in range(2,database_sheet.max_row):
        proj_id = database_sheet["A"+str(i)].value
        proj_year = str(proj_id)[:4]
        if int(proj_year) in years:#check if meets year criteria
            if database_sheet["D"+str(i)].value in projects:#check if meets project type criteria
                if database_sheet["F"+str(i)].value in clients:#check if meets client type criteria
                    if not loc_spec_var.get():#if we do NOT need to sort by location
                        #copy to excel sheet
                        write_sheet["A"+str(write_sheet_row)].value = database_sheet["A"+str(i)].value
                        write_sheet["B"+str(write_sheet_row)].value = database_sheet["B"+str(i)].value
                        write_sheet["C"+str(write_sheet_row)].value = database_sheet["C"+str(i)].value
                        write_sheet["D"+str(write_sheet_row)].value = database_sheet["D"+str(i)].value
                        write_sheet["E"+str(write_sheet_row)].value = database_sheet["E"+str(i)].value
                        write_sheet["F"+str(write_sheet_row)].value = database_sheet["F"+str(i)].value
                        write_sheet["G"+str(write_sheet_row)].value = database_sheet["G"+str(i)].value
                        write_sheet["H"+str(write_sheet_row)].value = database_sheet["H"+str(i)].value
                        write_sheet_row +=1
                    else:
                        #test to see if it is within radius
                        #distance = geodesic(tuple1lat,long,tuple2lat,long).miles
                        if database_sheet["G"+str(i)].value != None:
                            t_coords = (float(database_sheet["G"+str(i)].value),float(database_sheet["H"+str(i)].value))
                            distance = geodesic(coords,t_coords).miles
                            statusbar_text.set(distance)
                            if distance < radius:
                                write_sheet["A"+str(write_sheet_row)].value = database_sheet["A"+str(i)].value
                                write_sheet["B"+str(write_sheet_row)].value = database_sheet["B"+str(i)].value
                                write_sheet["C"+str(write_sheet_row)].value = database_sheet["C"+str(i)].value
                                write_sheet["D"+str(write_sheet_row)].value = database_sheet["D"+str(i)].value
                                write_sheet["E"+str(write_sheet_row)].value = database_sheet["E"+str(i)].value
                                write_sheet["F"+str(write_sheet_row)].value = database_sheet["F"+str(i)].value
                                write_sheet["G"+str(write_sheet_row)].value = database_sheet["G"+str(i)].value
                                write_sheet["H"+str(write_sheet_row)].value = database_sheet["H"+str(i)].value
                                write_sheet["k"+str(write_sheet_row)].value = distance
                                write_sheet_row +=1
    database_wb.remove(database_wb[DATABASE_SHEET_NAME])
    database_wb.save('RACEGIS_OUTPUT.xlsx')
    os.startfile("RACEGIS_OUTPUT.xlsx")


def all_client():
    if all_client_var.get():
        residential_var.set(1)
        government_var.set(1)
        myc_var.set(1) 
        util_var.set(1) 
        consultant_var.set(1)
        contractor_var.set(1)
        statusbar_text.set("Selecting all clients")
    else:
        residential_var.set(0)
        government_var.set(0)
        myc_var.set(0) 
        util_var.set(0) 
        consultant_var.set(0)
        contractor_var.set(0)
        statusbar_text.set("deselecting all clients")

def client():
    if residential_var.get() and \
         government_var.get() and \
         myc_var.get() and \
         util_var.get() and \
         consultant_var.get() and \
         contractor_var.get():
            all_client_var.set(1)
    else:
        all_client_var.set(0)


def all_proj():
    if all_proj_var.get():
        dredge_var.set(1)
        floodwave_var.set(1)
        hydro_var.set(1) 
        pierdock_var.set(1) 
        floodcontrol_var.set(1)
        upland_var.set(1)
        facilities_var.set(1)
        beach_var.set(1)
        construction_var.set(1)
        other_var.set(1)
        statusbar_text.set("Selecting all projects")

    else:
        dredge_var.set(0)
        floodwave_var.set(0)
        hydro_var.set(0) 
        pierdock_var.set(0) 
        floodcontrol_var.set(0)
        upland_var.set(0)
        facilities_var.set(0)
        beach_var.set(0)
        construction_var.set(0)
        other_var.set(0)
        statusbar_text.set("deselecting all projects")

    
def proj():
    '''
    Dummy function, the states will be checked later, all this does is set all on or off
    It checks if all the projects are checked, puts all on
    if all are not checked, leaves (or sets) it unchecked.
    '''
    if dredge_var.get() and \
         floodwave_var.get() and \
         hydro_var.get() and \
         pierdock_var.get() and \
         floodcontrol_var.get() and \
         upland_var.get() and \
         facilities_var.get() and \
         beach_var.get() and \
         construction_var.get() and \
         other_var.get():
            all_proj_var.set(1)
    else:
        all_proj_var.set(0)

def location_spec():
    '''
    This function checks the status of the 'specify location' check box
    is the box is checked (.get() returns true) then the state of all the location boxes is set to normal
    meaning that they are no longer greyed out.
    If the box is unchecked, they will all revert to being greyed out.
    '''
    if loc_spec_var.get():
        location_drop_down.config(state=tk.NORMAL)
        loc_entry.config(state=tk.NORMAL)
        loc_radius.config(state=tk.NORMAL)
        loc_radius_label.config(state=tk.NORMAL)
        starting_point_label.config(state=tk.NORMAL)


    else:
        location_drop_down.config(state=tk.DISABLED)    
        loc_entry.config(state=tk.DISABLED)
        loc_radius.config(state=tk.DISABLED)    
        loc_radius_label.config(state=tk.DISABLED)
        starting_point_label.config(state=tk.DISABLED)


      
def refresh_year_list(year):
    '''
    This function is called when the two year drop down menus are interacted with
    the first conditions catch certain behavior that might cause and error
    after that it calculates which years to have in the YEARS list which is a global variable which the passes_filter() function checks 
    '''
    if year_var_2.get() == "1996":
        #if the second year is set to 1996 (the first year, then the first one must be set to that as well
        year_var_1.set(year_var_2.get())
        statusbar_text.set("Setting start year to 1996")
    elif year_var_2.get() < year_var_1.get():
        #if the second (later year) is set to less than the previous year, set the first year to 1 year less so it works
        year_var_1.set(str(int(year_var_2.get())-1))
        statusbar_text.set("Start year cannot exceed end year, changing")
    statusbar_text.set("Time range changed")

    #Check to see if year_var_1 > year_var 2, change year_var 1 to be a year less, unless the year is 1996
    


#Create the window    
root = tk.Tk()
#Set the titlebar
root.title("RACE GIS v"+str(VERSION))

#YEARS
year_choices = []
year_var_1 = tk.StringVar()
year_var_1.set(str(now.year))
year_var_2 = tk.StringVar()
year_var_2.set(str(now.year))
for x in range(1996,now.year+1):
    year_choices.append(x)
years = ttk.Labelframe(root,text="Years")
years.grid(row=0, column=1,columnspan=2,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)
from_label = tk.Label(years, text="From")
from_label.grid(row=0,column=1)
year_var_1 = tk.StringVar()
year_var_1.set(str(now.year))
year_menu_1 = tk.OptionMenu(years,year_var_1, *year_choices,command=refresh_year_list)
year_menu_1.grid(row=0,column=2)
to_label = tk.Label(years, text="to")
to_label.grid(row=0,column=3)
year_var_2 = tk.StringVar()
year_var_2.set(str(now.year))
year_menu_2 = tk.OptionMenu(years,year_var_2, *year_choices,command=refresh_year_list)
year_menu_2.grid(row=0,column=4)

#PROJECTS
projects = ttk.Labelframe(root,text="Project Types")
projects.grid(row=1,rowspan=2,column=1,sticky=tk.N,ipadx=2,ipady=2,padx=2,pady=2)
all_proj_var = tk.IntVar(value=1)
all_proj_checkbox = tk.Checkbutton(projects, text="All",variable=all_proj_var,command=all_proj)
all_proj_checkbox.grid(row=0, column=1,sticky=tk.W)
dredge_var = tk.IntVar(value=1)
dredge_checkbox = tk.Checkbutton(projects, text="Dredging",variable=dredge_var,command=proj)
dredge_checkbox.grid(row=1, column=1,sticky=tk.W)
floodwave_var = tk.IntVar(value=1)
floodwave_checkbox = tk.Checkbutton(projects, text="Flood / Wave Study",variable=floodwave_var,command=proj)
floodwave_checkbox.grid(row=2, column=1,sticky=tk.W)
hydro_var = tk.IntVar(value=1)
hydro_checkbox = tk.Checkbutton(projects, text="Hydrographic Survey",variable=hydro_var,command=proj)
hydro_checkbox.grid(row=3, column=1,sticky=tk.W)
pierdock_var = tk.IntVar(value=1)
pierdock_checkbox = tk.Checkbutton(projects, text="Pier / Dock",variable=pierdock_var,command=proj)
pierdock_checkbox.grid(row=4, column=1,sticky=tk.W)
floodcontrol_var = tk.IntVar(value=1)
floodcontrol_checkbox = tk.Checkbutton(projects, text="Flood and erosion control structure",variable=floodcontrol_var,command=proj)
floodcontrol_checkbox.grid(row=5, column=1,sticky=tk.W)
upland_var = tk.IntVar(value=1)
upland_checkbox = tk.Checkbutton(projects, text="Upland Structure",variable=upland_var,command=proj)
upland_checkbox.grid(row=6, column=1,sticky=tk.W)
facilities_var = tk.IntVar(value=1)
facilities_checkbox = tk.Checkbutton(projects, text="Facilities Assessment",variable=facilities_var,command=proj)
facilities_checkbox.grid(row=7, column=1,sticky=tk.W)
beach_var = tk.IntVar(value=1)
beach_checkbox = tk.Checkbutton(projects, text="Beach / Dune / Living Shoreline",variable=beach_var,command=proj)
beach_checkbox.grid(row=8, column=1,sticky=tk.W)
construction_var = tk.IntVar(value=1)
construction_checkbox = tk.Checkbutton(projects, text="Construction Administration",variable=construction_var,command=proj)
construction_checkbox.grid(row=9, column=1,sticky=tk.W)
other_var = tk.IntVar(value=1)
other_checkbox = tk.Checkbutton(projects, text="Other",variable=other_var,command=proj)
other_checkbox.grid(row=10, column=1,sticky=tk.W)

#CLIENTS
clients = ttk.Labelframe(root,text="Client Types")
clients.grid(row=1,column=2,rowspan=2,sticky=tk.N+tk.S,ipadx=2,ipady=2,padx=2,pady=2)
all_client_var = tk.IntVar(value=1)
all_client_checkbox = tk.Checkbutton(clients, text="All",variable=all_client_var,command=all_client)
all_client_checkbox.grid(row=0, column=1,sticky=tk.W)
residential_var = tk.IntVar(value=1)
residential_checkbox = tk.Checkbutton(clients, text="Residential",variable=residential_var,command=client)
residential_checkbox.grid(row=1, column=1,sticky=tk.W)
government_var = tk.IntVar(value=1)
government_checkbox = tk.Checkbutton(clients, text="Goverment",variable=government_var,command=client)
government_checkbox.grid(row=2, column=1,sticky=tk.W)
myc_var = tk.IntVar(value=1)
myc_checkbox = tk.Checkbutton(clients, text="Marina / Yacht Club",variable=myc_var,command=client)
myc_checkbox.grid(row=3, column=1,sticky=tk.W)
util_var = tk.IntVar(value=1)
util_checkbox = tk.Checkbutton(clients, text="Utility / Industrial",variable=util_var,command=client)
util_checkbox.grid(row=4, column=1,sticky=tk.W)
consultant_var = tk.IntVar(value=1)
consultant_checkbox = tk.Checkbutton(clients, text="Consultant",variable=consultant_var,command=client)
consultant_checkbox.grid(row=5, column=1,sticky=tk.W)
contractor_var = tk.IntVar(value=1)
contractor_checkbox = tk.Checkbutton(clients, text="Contractor",variable=contractor_var,command=client)
contractor_checkbox.grid(row=6, column=1,sticky=tk.W)

#EXTRAS
extras = ttk.Labelframe(root,text="Extras")
extras.grid(row=2,column=2,sticky=tk.W+tk.E+tk.S,ipadx=2,ipady=2,padx=2,pady=2)
permit_var = tk.IntVar(value=0)
permit_checkbox = tk.Checkbutton(extras, text="Permit",variable=permit_var)
permit_checkbox.grid(row=0, column=1,sticky=tk.W)
permit_checkbox.config(state=tk.DISABLED)   #remove once added 
boring_var = tk.IntVar(value=0)
boring_checkbox = tk.Checkbutton(extras, text="Boring",variable=boring_var)
boring_checkbox.grid(row=1, column=1,sticky=tk.W)
boring_checkbox.config(state=tk.DISABLED)  #remove once added  

#LOCATION
location_by_options = ['By Address','By ID']
location = ttk.Labelframe(root,text="Location")
location.grid(row=3,column=1,sticky=tk.W+tk.E,columnspan=2,ipadx=2,ipady=2,padx=2,pady=2)
loc_spec_var = tk.IntVar(value=0)
location_spec = tk.Checkbutton(location,text="Specify Location", command=location_spec, var=loc_spec_var)
location_spec.grid(row=0,column=1)
location_var_1 = tk.StringVar()
location_var_1.set('By ID')
location_drop_down = tk.OptionMenu(location,location_var_1, *location_by_options)

location_drop_down.grid(column=2,row=0,sticky=tk.W+tk.E)
location_drop_down.config(state=tk.DISABLED)

starting_point_label = tk.Label(location,text="Starting Point")
starting_point_label.grid(column=1,row=2,sticky=tk.W+tk.E)
starting_point_label.config(state=tk.DISABLED)

loc_entry_var = tk.StringVar()
loc_entry = tk.Entry(location,textvariable=loc_entry_var)
loc_entry.config(state=tk.DISABLED)
loc_entry.config(font=("Consolas",10))

loc_entry.grid(column=2,row=2,sticky=tk.W+tk.E)

loc_radius_label = tk.Label(location,text="Radius (mi)")
loc_radius_label.grid(column=1,row=3,sticky=tk.W+tk.E)
loc_radius_label.config(state=tk.DISABLED)

loc_radius_var = tk.StringVar()
loc_radius = tk.Entry(location,textvariable=loc_radius_var)
loc_radius.config(state=tk.DISABLED)    
loc_radius_var.set("10")
loc_radius.grid(column=2,row=3)

#MAP
map_button = tk.Button(root,text='Google Earth',command=google_earth_button)
map_button.grid(row=4,column=1,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)

#SPREADSHEET
excel_button = tk.Button(root,text='Excel',command=excel_button)
excel_button.grid(row=4,column=2,sticky=tk.W+tk.E,ipadx=2,ipady=2,padx=2,pady=2)
#Status bar
statusbar_text = tk.StringVar()
status = tk.Label(root, text="",textvariable=statusbar_text, bd=1, relief=tk.SUNKEN, anchor=tk.W,font=("Consolas", 8))
status.grid(row=5,column=1,columnspan=2,sticky=tk.W+tk.E)

startup_database_check()
root.mainloop()
