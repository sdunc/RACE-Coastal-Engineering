#Stephen Duncanson
#For: RACE Coastal
#Written: July 11th 2019

import os, sys
import openpyxl

def main():
    file1 = open("all_projects.txt","w")
    file1.write('project,id,type,address')
    file1.write('\n')
    starting_row = 2
    ending_row = 497
    database = "Database.xlsx"
    print("opening file:",database)
    wb = openpyxl.load_workbook(database)                 #open a workbook
    sheet = wb['Database']#assign the sheet variable to a certain sheet
    for i in range(starting_row,ending_row):
        file1.write(str(sheet.cell(row=i,column=2).value))
        file1.write(',')#delimiter, I think thats what its called
        file1.write(str(sheet.cell(row=i,column=1).value))
        file1.write(',')#delimiter, I think thats what its called
        file1.write(str(sheet.cell(row=i,column=4).value))
        file1.write(',')#delimiter, I think thats what its called
        file1.write(str(sheet.cell(row=i,column=3).value))
        file1.write('\n')#delimiter, I think thats what its called
    file1.close()





if __name__ == '__main__': 
    main() 
